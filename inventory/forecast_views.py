from datetime import datetime
import os
from django.core.files.storage import FileSystemStorage
from django.contrib import messages
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from .utils import forecast_usage_from_excel, is_clerk, is_hod, is_faculty 
from inventory.models import ItemRequest, InventoryItem
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def inventory_forecast_view(request):
    forecast_results = None
    forecast_year = [ datetime.now().year + 1, datetime.now().year + 2 ] 
    year = datetime.now().year + 1  
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        year = request.POST.get('forecast_year')
        if uploaded_file.name.endswith('.xlsx'):
            fs = FileSystemStorage()
            filename = fs.save(uploaded_file.name, uploaded_file)
            file_path = fs.path(filename)

            try:
                
                forecast_results = forecast_usage_from_excel(file_path, int(year))

            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
            finally:
                os.remove(file_path)
        else:
            messages.error(request, "Please upload a .xlsx file only.")

    return render(request, 'forecast/forecast_page.html', {
        'forecast_data': forecast_results,
        'forecast_year': forecast_year,
        'selected_year': year
    })


@login_required
@user_passes_test(is_clerk or is_hod)
def predict_usage(request, year):
    forecast_data = {}
    try:
        year = int(year)
    except ValueError:
        return JsonResponse({'error': 'Invalid year format'}, status=400)
    forecast_year = year 

    excel_path = os.path.join('media', 'stationery_usage.xlsx') 
    try:
        forecast_data = forecast_usage_from_excel(excel_path, forecast_year)
    except Exception as e:
        forecast_data = {'error': str(e)}

    return render(request, 'predict_usage.html', {
        'forecast_data': forecast_data,
        'forecast_year': forecast_year
    })

import pandas as pd
def forecast_inventory_usage(request, forecast_years=1):
    issued_requests = ItemRequest.objects.filter(status='issued', item__item_type='consumable').order_by('request_date')
    if not issued_requests.exists():
        return render(request, 'predict_usage2.html', {
            'error': 'No issued item requests found in the database.',
            'forecast_data': {},
            'past_usage': {},
            'current_stock': {},
            'need_to_order': {},
            'forecast_year': datetime.now().year + 1,
            'years_to_show': []
        })

    current_year = datetime.now().year
    years_to_show = list(range(current_year - 4, current_year))  # Last 4 years excluding current

    items = InventoryItem.objects.filter(item_type='consumable')
    usage_data = {item.name: {year: 0 for year in years_to_show} for item in items}
    current_stock = {item.name: item.quantity for item in items}  # Current available quantity

    for item in items:
        item_requests = issued_requests.filter(item=item)
        for req in item_requests:
            year = req.request_date.year
            if year in usage_data[item.name]:
                usage_data[item.name][year] += req.quantity

    forecast_result = {}
    need_to_order = {}

    for item_name, yearly_data in usage_data.items():
        sorted_years = sorted(yearly_data.keys())
        values = [yearly_data[year] for year in sorted_years]

        idx = pd.PeriodIndex(sorted_years, freq='Y')
        series = pd.Series(values, index=idx)

        if series.sum() == 0 or series.count() < 3:
            forecast_result[item_name] = "Not enough data"
            need_to_order[item_name] = "-"
            continue

        try:
            model = ExponentialSmoothing(series, trend='add', seasonal=None)
            fit = model.fit()
            forecast = fit.forecast(1)  

            forecast_year = str(current_year + 1)
            forecast_value = int(round(forecast.iloc[0]))  
            forecast_result[item_name] = {forecast_year: forecast_value}

            available_qty = current_stock.get(item_name, 0)
            need_qty = forecast_value - available_qty
            need_to_order[item_name] = need_qty if need_qty > 0 else 0

        except Exception as e:
            forecast_result[item_name] = f"Forecasting error: {str(e)}"
            need_to_order[item_name] = "-"

    forecast_years_used = [str(current_year + 1)]

    return render(request, 'predict_usage2.html', {
        'forecast_data': forecast_result,
        'past_usage': usage_data,
        'current_stock': current_stock,
        'need_to_order': need_to_order,
        'forecast_year': current_year + 1,
        'years_to_show': years_to_show,
        'forecast_years_used': forecast_years_used,
    })

@csrf_exempt
def generate_forecast_excel_report(request):
    if request.method == 'POST':
        items = InventoryItem.objects.filter(item_type='consumable')
        current_stock = {item.name: item.quantity for item in items}

        item_data = []
        for item in items:
            field_name = f"order_qty_{item.name.lower().replace(' ', '-').replace('/', '-')}"

            order_qty = request.POST.get(field_name)
            try:
                order_qty = int(order_qty)
            except:
                order_qty = 0

            forecast_year = datetime.now().year + 1
            forecasted_qty = order_qty + current_stock.get(item.name, 0)
            prev_qty = forecasted_qty - order_qty

            item_data.append({
                "name": item.name,
                "sanctioned": forecasted_qty,
                "prev_qty": prev_qty,
                "prev_date": "22-07-2024",  # Replace with actual date 
                "balance": current_stock.get(item.name, 0),
                "required": order_qty,
            })

        # Generate Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stationery Forecast Report"

        headers = ["Sr. No.", "Stationery Items", "Sanctioned Qty", "Previous received qty. & Date", "Balance Qty.", "Current required Qty."]
        ws.append(headers)  

        thin = Side(border_style="thin", color="000000")

        for i, item in enumerate(item_data, start=1):
            row = [
                i,
                item["name"],
                item["sanctioned"],
                f"{item['prev_qty']} ({item['prev_date']})",
                item["balance"],
                item["required"]
            ]
            ws.append(row)

            for col in range(1, 7):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"stationery_forecast_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    return HttpResponse("Invalid request", status=400)

